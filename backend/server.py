from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import re
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List
import uuid
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

EXCEL_FILE = ROOT_DIR / 'contacts.xlsx'

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")


# --- Excel helpers ---

def _get_or_create_workbook():
    if EXCEL_FILE.exists():
        return load_workbook(EXCEL_FILE)
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["Sr.No", "Name", "Contact", "Gmail", "Message"])
    wb.save(EXCEL_FILE)
    return wb

def _phone_exists_in_excel(phone: str) -> bool:
    if not EXCEL_FILE.exists():
        return False
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        if row[0] and str(row[0]).strip() == phone.strip():
            return True
    return False

def _save_contact_to_excel(name: str, contact: str, gmail: str, message: str):
    wb = _get_or_create_workbook()
    ws = wb.active
    sr_no = ws.max_row
    ws.append([sr_no, name, contact, gmail, message])
    wb.save(EXCEL_FILE)

def _validate_phone(phone: str) -> bool:
    cleaned = re.sub(r'[\s\-\(\)\+]', '', phone)
    return bool(re.match(r'^\d{10}$', cleaned))


# --- Models ---

class ContactCreate(BaseModel):
    name: str
    email: str
    phone: str
    message: str

class EmailSubscription(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class EmailCreate(BaseModel):
    email: str

class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str


# --- Routes ---

@api_router.get("/")
async def root():
    return {"message": "GamLens API"}

@api_router.post("/contact")
async def create_contact(input: ContactCreate):
    if not input.phone.strip():
        raise HTTPException(status_code=400, detail="Phone number is required")
    if not _validate_phone(input.phone):
        raise HTTPException(status_code=400, detail="Please enter a valid 10 digit phone number")
    if _phone_exists_in_excel(input.phone):
        raise HTTPException(status_code=409, detail="This phone number is already registered")
    _save_contact_to_excel(name=input.name, contact=input.phone, gmail=input.email, message=input.message)
    return {"status": "success", "message": "Registration successful"}

@api_router.get("/download-contacts")
async def download_contacts():
    if not EXCEL_FILE.exists():
        raise HTTPException(status_code=404, detail="No contacts file found")
    return FileResponse(path=str(EXCEL_FILE), filename="contacts.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@api_router.post("/subscribe", response_model=EmailSubscription)
async def subscribe_email(input: EmailCreate):
    existing = await db.subscribers.find_one({"email": input.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already subscribed")
    subscription = EmailSubscription(**input.model_dump())
    doc = subscription.model_dump()
    await db.subscribers.insert_one(doc)
    doc.pop('_id', None)
    return subscription

@api_router.get("/subscribers", response_model=List[EmailSubscription])
async def get_subscribers():
    subs = await db.subscribers.find({}, {"_id": 0}).to_list(1000)
    return subs

@api_router.get("/notifications")
async def get_notifications():
    notifs = await db.notifications.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return notifs

@api_router.put("/notifications/{notif_id}/read")
async def mark_notification_read(notif_id: str):
    result = await db.notifications.update_one({"id": notif_id}, {"$set": {"read": True}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"status": "ok"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_obj = StatusCheck(**input.model_dump())
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    await db.status_checks.insert_one(doc)
    doc.pop('_id', None)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    return status_checks

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
